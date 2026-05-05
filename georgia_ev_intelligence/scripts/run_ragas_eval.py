"""
scripts/run_ragas_eval.py — Phase 4 RAGAS Evaluation with checkpoint/resume.

Usage:
  venv\\Scripts\\python scripts\\run_ragas_eval.py --questions 50
  venv\\Scripts\\python scripts\\run_ragas_eval.py --resume   # resume from checkpoint
"""
from __future__ import annotations

import argparse
import asyncio
import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from statistics import mean, pstdev

import httpx
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from phase4_agent.pipeline import EVAgent
from shared.config import Config
from shared.logger import get_logger

logger = get_logger("ragas_eval")

PROGRESS_FILE = ROOT / "outputs" / "progress" / "phase4_eval_progress.jsonl"
ANSWERS_MD    = ROOT / "outputs" / "progress" / "phase4_eval_answers.md"


def _sanitize_model_name(name: str) -> str:
    """Make model names filesystem-safe while keeping them recognizable."""
    return re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("._-") or "unknown-model"


def _default_report_path() -> Path:
    cfg = Config.get()
    embed_model = _sanitize_model_name(cfg.ollama_embed_model)
    generation_model = _sanitize_model_name(cfg.ollama_llm_model)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"phase4_ragas_{embed_model}_{generation_model}_{timestamp}.xlsx"
    return ROOT / "outputs" / "ragas_reports" / filename

# ── Metric config (matches ev_data_LLM_comparsions/config/config.yaml) ────────
WEIGHTS = {
    "faithfulness":       0.25,
    "answer_relevancy":   0.20,
    "context_precision":  0.20,
    "context_recall":     0.20,
    "answer_correctness": 0.15,
}

DEFINITIONS = {
    "faithfulness":
        "Every claim in the generated answer is grounded in the retrieved context. "
        "Score 1.0 only if NO statement goes beyond the retrieved data.",
    "answer_relevancy":
        "The answer directly and completely addresses the user question. "
        "Score 1.0 if fully on-topic and nothing important is missing.",
    "context_precision":
        "The retrieved context chunks are relevant and useful for the question. "
        "Score 1.0 if all retrieved data is on-topic (no noise).",
    "context_recall":
        "The retrieved context covers all key facts present in the golden answer. "
        "Score 1.0 if every golden fact can be found in context.",
    "answer_correctness":
        "The generated answer is factually aligned with the human-validated golden answer. "
        "Score 1.0 if all facts match.",
}

# ── 7 smoke questions (extend to 50 by passing your full list) ─────────────────
SMOKE_QUESTIONS: list[dict] = [
    {
        "id": "Q1", "category": "AGGREGATE",
        "question": "Which county has the highest total employment among Tier 1 suppliers only?",
        "golden": "Troup County has the highest total employment among Tier 1 suppliers.",
    },
    {
        "id": "Q2", "category": "RISK",
        "question": "Which EV supply chain roles in Georgia have only one supplier, making them a single point of failure?",
        "golden": "Georgia has multiple single-supplier roles including Charging Infrastructure (Morgan Corp.), Power Electronics (GSC Steel Stamping), and Materials (Haering Precision USA LP).",
    },
    {
        "id": "Q3", "category": "CYPHER-TIER",
        "question": "Which Georgia companies are classified under Battery Cell or Battery Pack roles, and what tier is each?",
        "golden": "6 companies: Hitachi Astemo Americas (Tier 1/2, Battery Cell), Hollingsworth & Vose (Tier 1/2, Battery Pack), Honda Development & Manufacturing (Tier 1/2, Battery Cell), Hyundai Motor Group (Tier 1/2, Battery Pack), F&P Georgia Manufacturing (Tier 1/2, Battery Pack), IMMI (Tier 1/2, Battery Pack).",
    },
    {
        "id": "Q4", "category": "CYPHER-OEM",
        "question": "Show the full supplier network linked to Rivian Automotive in Georgia, broken down by tier and EV Supply Chain Role.",
        "golden": "Georgia suppliers linked to Rivian include GSC Steel Stamping (Tier 2/3), Duckyang (Tier 2/3), Enchem America (Tier 2/3), Hyundai Transys Georgia Powertrain (Tier 1/2), and Remark International LLC (Tier 1).",
    },
    {
        "id": "Q5", "category": "CYPHER-PROD",
        "question": "Find Georgia-based companies that manufacture copper foil or electrodeposited materials for EV battery current collectors.",
        "golden": "Duckyang in Jackson County manufactures electrodeposited copper foil for EV battery current collectors.",
    },
    {
        "id": "Q6", "category": "CYPHER-LOC",
        "question": "In Gwinnett County, which company has the highest employment and what is its EV Supply Chain Role?",
        "golden": "SungEel Recycling Park Georgia has the highest employment in Gwinnett County with 650 employees. Its EV Supply Chain Role is Materials.",
    },
    {
        "id": "Q7", "category": "CYPHER-FAC",
        "question": "Which Georgia companies operate R&D facilities focused on EV technology?",
        "golden": "Racemark International LLC (Jones County, Tier 1) operates R&D facilities focused on EV technology.",
    },
]


FIFTY_QUESTIONS: list[dict] = [
    {"id": "Q1", "category": "Supply Chain Mapping & Visibility", "question": "Show all \"Tier 1/2\" suppliers in Georgia, list their EV Supply Chain Role and Product / Service.", "golden": "There are 18 Tier 1/2 companies in Georgia (using Updated Location).\nF&P Georgia Manufacturing [Tier 1/2] | Role: Battery Pack | Product: Lithium-ion battery recycler and raw materials provider\nFouts Brothers Fire Equipment [Tier 1/2] | Role: General Automotive | Product: Automotive body parts and electronics Cylinder-head and specialty gaskets, housing modules and shielding systems for engines, transmissions, exhaust systems and auxiliary units\nHitachi Astemo [Tier 1/2] | Role: General Automotive | Product: Car audio systems\nHitachi Astemo Americas Inc. [Tier 1/2] | Role: Battery Cell | Product: Battery cells for electric mobility\nHollingsworth & Vose Co. [Tier 1/2] | Role: Battery Pack | Product: Lithium-ion battery materials\nHonda Development & Manufacturing [Tier 1/2] | Role: Battery Cell | Product: Battery cells for electric mobility\nHwashin [Tier 1/2] | Role: General Automotive | Product: Modules and storage systems\nHyundai & LG Energy Solution (LGES) [Tier 1/2] | Role: General Automotive | Product: Storage batteries\nHyundai Industrial Co. [Tier 1/2] | Role: General Automotive | Product: Automotive batteries\nHyundai MOBIS (Georgia) [Tier 1/2] | Role: General Automotive | Product: Capacitors, electronic automotive components\nHyundai Motor Group [Tier 1/2] | Role: Battery Pack | Product: Battery parts for electric vehicles\nHyundai Transys Georgia Powertrain [Tier 1/2] | Role: Thermal Management | Product: Electrical heaters, control units, and actuators\nHyundai Transys Georgia Seating Systems [Tier 1/2] | Role: General Automotive | Product: Automotive electronics, resonators, capacitors, resistors, electronics and electronic parts\nIMMI [Tier 1/2] | Role: Battery Pack | Product: Battery electrolyte\nIMS Gear Georgia Inc. [Tier 1/2] | Role: General Automotive | Product: Car audio systems\nInalfa Roof Systems Inc. [Tier 1/2] | Role: General Automotive | Product: Automotive electronic safety systems\nJAC Products Inc. [Tier 1/2] | Role: General Automotive | Product: Switches, resistors and related products\nJefferson Southern Corp. [Tier 1/2] | Role: General Automotive | Product: High current switches and resistors for automotive HVAC systems"},
    {"id": "Q2", "category": "Supply Chain Mapping & Visibility", "question": "Which Georgia companies are classified under Battery Cell or Battery Pack roles, and what tier is each assigned?", "golden": "There are 6 companies found with Battery Cell or Battery Pack roles.    \n\nF&P Georgia Manufacturing [Tier 1/2] | Role: Battery Pack \nHitachi Astemo Americas Inc. [Tier 1/2] | Role: Battery Cell \nHollingsworth & Vose Co. [Tier 1/2] | Role: Battery Pack \nHonda Development & Manufacturing [Tier 1/2] | Role: Battery Cell \nHyundai Motor Group [Tier 1/2] | Role: Battery Pack \nIMMI [Tier 1/2] | Role: Battery Pack"},
    {"id": "Q3", "category": "Supply Chain Mapping & Visibility", "question": "Map all Thermal Management suppliers in Georgia and show which Primary OEMs they are linked to.", "golden": "There are 5 Thermal Management companies in Georgia.\nZF Gainesville LLC | OEMs: Multiple OEMs\nFreudenberg-NOK | OEMs: Multiple OEMs\nHyundai Transys Georgia Powertrain | OEMs: Hyundai Kia Rivian\nNovelis Inc. | OEMs: Multiple OEMs\nPeerless-Winsmith Inc. | OEMs: Multiple OEMs"},
    {"id": "Q4", "category": "Supply Chain Mapping & Visibility", "question": "List every Georgia company classified under Power Electronics or Charging Infrastructure, along with their Employment size.", "golden": "There are 4 companies strictly classified under Power Electronics or Charging Infrastructure:\nGSC Steel Stamping LLC | Role: Power Electronics | Employment: 350\nMorgan Corp. | Role: Charging Infrastructure | Employment: 320\nYazaki North America | Role: Power electronics, sensors, and EV systems | Employment: 230000\nZF Gainesville LLC has a combined role (EV thermal management and power electronics) | Employment: 17500"},
    {"id": "Q5", "category": "Supply Chain Mapping & Visibility", "question": "Which companies are classified as Direct Manufacturer, and what EV Supply Chain Roles do they cover?", "golden": "Original Equipment Manufacturer companies (11):\nKia Georgia Inc. | EV Supply Chain Role: Vehicle Assembly\nMinebea AccessSolutions USA Inc. | EV Supply Chain Role: General Automotive\nSuperior Essex Inc. | EV Supply Chain Role: Vehicle Assembly\nSuzuki Manufacturing of America Corp. | EV Supply Chain Role: Vehicle Assembly\nTCI Powder Coatings | EV Supply Chain Role: Vehicle Assembly\nTDK Components USA Inc. | EV Supply Chain Role: Vehicle Assembly\nTE Connectivity | EV Supply Chain Role: Vehicle Assembly\nTeklas USA | EV Supply Chain Role: Vehicle Assembly\nTextron Specialized Vehicles | EV Supply Chain Role: Vehicle Assembly\nThermal Ceramics Inc. | EV Supply Chain Role: Vehicle Assembly\nThomson Plastics Inc. | EV Supply Chain Role: Vehicle Assembly"},
    {"id": "Q6", "category": "Supply Chain Mapping & Visibility", "question": "What locations does Novelis Inc. operate in, and what primary facility types are associated with each location?", "golden": "Novelis Inc. has three operating entries in Georgia, and all three are located in Atlanta, Fulton County, with each one classified as a Manufacturing Plant as its primary facility type."},
    {"id": "Q7", "category": "Supply Chain Mapping & Visibility", "question": "In Gwinnett County, which company has the highest Employment and what is its EV Supply Chain Role?", "golden": "WIKA USA has the highest employment in Gwinnett County (Employment: 250000; Role: HV and LV wiring harnesses for EVs and ICE vehicles)."},
    {"id": "Q8", "category": "Supply Chain Mapping & Visibility", "question": "Which county have the highest total Employment among Tier 1 suppliers only?", "golden": "Troup County has the highest total Employment among Tier 1 suppliers with a total of 2,435 employees"},
    {"id": "Q9", "category": "Supply Chain Mapping & Visibility", "question": "Which county has the highest total employment across all companies, and what is the combined employment in that county?", "golden": "Gwinnett County has the highest total employment across all companies: 253022."},
    {"id": "Q10", "category": "Supply Chain Mapping & Visibility", "question": "Identify all Vehicle Assembly facilities in Georgia and list the corresponding Primary OEM associated with each facility.", "golden": "There are 10 Vehicle Assembly OEMs along with their Primary OEMs are:\nKia Georgia Inc. | Primary OEMs: Club Car LLC\nSuperior Essex Inc. | Primary OEMs: Hyundai Motor Group\nSuzuki Manufacturing of America Corp. | Primary OEMs: Rivian Automotive\nTCI Powder Coatings | Primary OEMs: Kia Georgia Inc.\nTDK Components USA Inc. | Primary OEMs: Mercedes-Benz USA LLC\nTE Connectivity | Primary OEMs: Blue Bird Corp.\nTeklas USA | Primary OEMs: Yamaha Motor Manufacturing Corp.\nTextron Specialized Vehicles | Primary OEMs: Textron Specialized Vehicles\nThermal Ceramics Inc. | Primary OEMs: SK Battery America\nThomson Plastics Inc. | Primary OEMs: Archer Aviation Inc."},
    {"id": "Q11", "category": "Supply Chain Mapping & Visibility", "question": "Identify the primary products and services associated with Sewon America Inc. across its different operational sites in Georgia.", "golden": "Sewon America Inc. has three operating entries in Georgia and all three are located in LaGrange, Troup County, Georgia.\nSewon America Inc. | Product: Motor vehicle engines and parts\nSewon America Inc. | Product: Fire truck bodies\nSewon America Inc. | Product: Automotive aftermarket parts"},
    {"id": "Q12", "category": "Supply Chain Mapping & Visibility", "question": "List all Tier 2/3 companies in Georgia with primary involvement in the electric vehicle or battery supply chain and specify their respective roles.", "golden": "There are 3 companies with primary involvement in the electric vehicle or battery supply chain. \n\nDuckyang | Role: General Automotive \nGSC Steel Stamping LLC | Role: Power Electronics \nEnchem America Inc. | Role: General Automotive"},
    {"id": "Q13", "category": "Supply Chain Mapping & Visibility", "question": "Show the full supplier network linked to Rivian Automotive in Georgia, broken down by tier and EV Supply Chain Role.", "golden": "Rivian-linked suppliers in GA (by Primary OEMs column):\nDuckyang [Tier 2/3] | Role: General Automotive | Primary OEMs: Hyundai Kia Rivian\nEnchem America Inc. [Tier 2/3] | Role: General Automotive | Primary OEMs: Hyundai Kia Rivian\nGSC Steel Stamping LLC [Tier 2/3] | Role: Power Electronics | Primary OEMs: Hyundai Kia Rivian\nHyundai Transys Georgia Powertrain [Tier 1/2] | Role: Thermal Management | Primary OEMs: Hyundai Kia Rivian\nRacemark International LLC [Tier 1] | Role: General Automotive | Primary OEMs: Hyundai Kia Rivian\nSuzuki Manufacturing of America Corp. [OEM] | Role: Vehicle Assembly | Primary OEMs: Rivian Automotive"},
    {"id": "Q14", "category": "Supply Chain Mapping & Visibility", "question": "Which Georgia companies produce battery materials such as anodes, cathodes, electrolytes, or copper foil, and what tier are they classified as?", "golden": "There are 4 Georgia companies producing battery materials such as anodes, cathodes, electrolytes, or copper foil. \nDuckyang [Tier 2/3] | Produce: High-quality electrodeposited (ED) copper foil for electric vehicles\nF&P Georgia Manufacturing [Tier 1/2] | Produce: Lithium-ion battery recycler and raw materials provider\nHollingsworth & Vose Co. [Tier 1/2] | Produce: Lithium-ion battery materials\nIMMI [Tier 1/2]  | Produce: Battery electrolyte"},
    {"id": "Q15", "category": "Supply Chain Mapping & Visibility", "question": "Identify all Georgia companies with an EV Supply Chain Role related to wiring harnesses and show their Primary OEMs.", "golden": "There are 2 Georgia companies with an EV Supply Chain Role related to wiring harnesses:\nWIKA USA | Primary OEMs: n/a\nWoodbridge Foam Corp. | Primary OEMs: n/a"},
    {"id": "Q16", "category": "Supplier Discovery & Matchmaking", "question": "Find Georgia-based Tier 1 or Tier 1/2 suppliers capable of producing battery electrolytes or lithium-ion battery materials with existing OEM contracts.", "golden": "The Georgia-based suppliers that match battery electrolytes or lithium-ion battery materials and also have existing OEM contracts are:\nHollingsworth & Vose Co. [Tier 1/2] | OEMs: Hyundai Kia | Produces: Lithium-ion battery materials\nIMMI [Tier 1/2] | OEMs: Hyundai Kia | Produces: Battery electrolyte"},
    {"id": "Q17", "category": "Supplier Discovery & Matchmaking", "question": "Which Georgia companies manufacture high-voltage wiring harnesses or EV electrical distribution components suitable for BEV platforms?", "golden": "There are 2 Georgia companies manufacturing high-voltage wiring harnesses or EV electrical distribution components suitable for BEV platforms\n\nWIKA USA | Role: HV and LV wiring harnesses for EVs and ICE vehicles | Product: Vehicle power & data solutions; wiring harnesses and connectors | Ev Relevant: Yes\nWoodbridge Foam Corp. | Role: EV wiring harnesses and power distribution | Product: Automotive wiring harnesses & electrical distribution systems | Ev Relevant: Yes"},
    {"id": "Q18", "category": "Supplier Discovery & Matchmaking", "question": "Identify Georgia Tier 2/3 companies in the Electronic and Electrical Equipment industry group that could be upgraded to supply EV power electronics.", "golden": "There are no Georgia Tier 2/3 companies listed under the Electronic and Other Electrical Equipment and Components industry group. So none can be upgraded to supply EV power electronics based on the provided evidence."},
    {"id": "Q19", "category": "Supplier Discovery & Matchmaking", "question": "Find Georgia-based companies that manufacture copper foil or electrodeposited materials suitable for EV battery current collectors.", "golden": "There is 1 company in Georgia specifically identified for the manufacture of materials suitable for EV battery current collectors:\nDuckyang | Produce: High-quality electrodeposited (ED) copper foil for electric vehicles"},
    {"id": "Q20", "category": "Supplier Discovery & Matchmaking", "question": "Which Georgia Tier 1/2 companies produce engineered plastics, polymers, or composite materials applicable to EV structural or thermal components?", "golden": "No Georgia Tier 1/2 companies are explicitly identified as producing engineered plastics, polymers, or composite materials applicable to EV structural or thermal components."},
    {"id": "Q21", "category": "Supplier Discovery & Matchmaking", "question": "Find Georgia suppliers with existing Hyundai Kia contracts that could be expanded to support Hyundai Metaplant's EV battery production ramp-up.", "golden": "There are 11 companies with existing Hyundai Kia contracts that could be expanded to support Hyundai Metaplant's EV battery production ramp-up.  \n\nF&P Georgia Manufacturing \nDuckyang \nGSC Steel Stamping LLC \nHitachi Astemo Americas Inc. \nHollingsworth & Vose Co. \nHonda Development & Manufacturing \nHyundai Motor Group \nHyundai Transys Georgia Powertrain \nIMMI \nEnchem America Inc. \nRacemark International LLC"},
    {"id": "Q22", "category": "Supplier Discovery & Matchmaking", "question": "Identify Georgia companies producing DC-to-DC converters, capacitors, or power electronics components relevant to EV drivetrains and what tier is each assigned?", "golden": "There are 3 Georgia companies producing DC-to-DC converters, capacitors, or power electronics components relevant to EV drivetrains. \n\nGSC Steel Stamping LLC [Tier 2/3] \nHyundai MOBIS (Georgia) [Tier 1/2] \nHyundai Transys Georgia Seating Systems [Tier 1/2]"},
    {"id": "Q23", "category": "Supplier Discovery & Matchmaking", "question": "Which Georgia companies provide powder coating-related products or services, and what tier are they classified under?", "golden": "There is only 1 Georgia company that provides Powder coating-related products/services:\nArcher Aviation Inc. [Tier 2/3] | Produces: Powder coatings for automotive and other industries"},
    {"id": "Q24", "category": "Supplier Discovery & Matchmaking", "question": "Which Georgia companies manufacture battery parts or enclosure systems and are classified as Tier 1/2, making them ready for direct OEM engagement and show which Primary OEMs they are linked to.", "golden": "There are 6 Georgia companies manufacturing battery parts or enclosure systems and are classified as Tier 1/2, making them ready for direct OEM engagement. \n\nF&P Georgia Manufacturing | OEMs: Hyundai Kia \nHitachi Astemo Americas Inc. | OEMs: Hyundai Kia \nHollingsworth & Vose Co. | OEMs: Hyundai Kia \nHonda Development & Manufacturing | OEMs: Hyundai Kia \nHyundai Motor Group | OEMs: Hyundai Kia \nIMMI | OEMs: Hyundai Kia"},
    {"id": "Q25", "category": "Supplier Discovery & Matchmaking", "question": "Find Tier 2/3 Georgia-based suppliers with employment over 300 that are classified as General Automotive but produce components transferable to EV platforms.", "golden": "There are 8 Tier 2/3 Georgia-based suppliers with employment over 300 that are classified as General Automotive but produce components transferable to EV platforms. \n\nArising Industries Inc. | Employment: 582 \nDinex Emissions Inc. | Employment: 500 \nACM Georgia LLC | Employment: 400 \nDongwon Autopart Technology Georgia LLC | Employment: 315 \nFOX Factory | Employment: 345 \nAVS | Employment: 310 \nFlambeau Inc. | Employment: 700\nGrudem | Employment: 807"},
    {"id": "Q26", "category": "Supplier Discovery & Matchmaking", "question": "Identify Georgia Tier 2/3 companies in the Chemicals and Allied Products industry group and list their products.", "golden": "There are 2 Georgia Tier 2/3 companies in Chemicals and Allied Products industry group.\n\nArcher Aviation Inc. | Product: Powder coatings for automotive and other industries \nArising Industries Inc. | Product: Rubber powders for tires"},
    {"id": "Q27", "category": "Supply Chain Risk & Resilience", "question": "Which EV Supply Chain Roles in Georgia are served by only a single company, creating a single-point-of-failure risk for the state's EV ecosystem?", "golden": "There are 28 Supply Chain Roles in Georgia served by only a single company, creating a single-point-of-failure risk for the state's EV ecosystem  \n\nRole: Advanced electrical architecture for EVs | Only company: Yamaha Motor Manufacturing Corp.\nRole: Body and chassis components for EV and ICE OEMs | Only company: Vanguard National Trailer Corp.\nRole: Charging Infrastructure | Only company: Morgan Corp.\nRole: EV and ICE component manufacturing | Only company: Vista Metals Corp.\nRole: EV and ICE powertrain and HVAC components | Only company: Voestalpine Automotive Body Parts Inc.\nRole: EV body, powertrain, and thermal components | Only company: ZF Gainesville LLC\nRole: EV electrical distribution and interior systems | Only company: Yachiyo Manufacturing of America LLC\nRole: EV sensors, braking, and control systems | Only company: YKK USA Inc.\nRole: EV thermal management and power electronics | Only company: ZF Gainesville LLC\nRole: EV thermal systems and electronics | Only company: ZF Gainesville LLC\nRole: EV wiring harnesses and power distribution | Only company: Woodbridge Foam Corp.\nRole: HV and LV wiring harnesses for EVs and ICE vehicles | Only company: WIKA USA\nRole: High\u2011voltage EV connectors and electronics | Only company: Woory Industrial Co.\nRole: Interior and exterior plastic parts | Only company: Wabash National Corp.\nRole: Interior trim and textile components | Only company: Wheelabrator Group Inc.\nRole: Machined components supporting EV and ICE vehicles | Only company: Trenton Pressing Inc.\nRole: OEM corporate and engineering footprint (electrification) | Only company: Tie Down Engineering\nRole: OEM corporate footprint and supplier integration | Only company: Trenton Pressing\nRole: OEM corporate footprint influencing EV strategy | Only company: TI Fluid Systems\nRole: OEM parent group footprint (EV + HD electrification) | Only company: TN Americas Holding Inc.\nRole: OEM parent group footprint (electric truck strategy) | Only company: Toyota Industries Group (TACG-TICA)\nRole: Plastic components for EV and ICE vehicles | Only company: Volvo Cars USA\nRole: Power Electronics | Only company: GSC Steel Stamping LLC\nRole: Power electronics, sensors, and EV systems | Only company: Yazaki North America\nRole: Stamped and welded assemblies for OEMs | Only company: Vernay\nRole: Stamped metal components for OEMs | Only company: Volvo Group North America\nRole: Tier 1 automotive components | Only company: Valeo\nRole: Vehicle safety systems OEM (EV + ICE) \u2192 Only company: TI Fluid Systems"},
    {"id": "Q28", "category": "Supply Chain Risk & Resilience", "question": "Which Georgia Battery Cell or Battery Pack suppliers are sole-sourced by a specific OEM, indicating high dependency risk?", "golden": "There are 6 Georgia Battery Cell or Battery Pack suppliers that are sole-sourced by a specific OEM, indicating high dependency risk.\n  \u2022 F&P Georgia Manufacturing | Role: Battery Pack | Primary OEM: Hyundai Kia\n  \u2022 Hitachi Astemo Americas Inc. | Role: Battery Cell | Primary OEM: Hyundai Kia\n  \u2022 Hollingsworth & Vose Co. | Role: Battery Pack | Primary OEM: Hyundai Kia\n  \u2022 Honda Development & Manufacturing | Role: Battery Cell | Primary OEM: Hyundai Kia\n  \u2022 Hyundai Motor Group | Role: Battery Pack | Primary OEM: Hyundai Kia\n  \u2022 IMMI | Role: Battery Pack | Primary OEM: Hyundai Kia"},
    {"id": "Q29", "category": "Supply Chain Risk & Resilience", "question": "For Hyundai Metaplant, how many of its Georgia-based EV component suppliers have fewer than 200 employees, flagging potential capacity fragility?", "golden": "For Hyundai Metaplant, there are 6 Georgia-based EV component suppliers with less than 200 employees, flagging potential capacity fragility.\nEnchem America Inc. | Employment: 155\nF&P Georgia Manufacturing | Employment: 104\nHyundai Motor Group | Employment: 164\nHyundai Transys Georgia Powertrain | Employment: 130\nIMMI | Employment: 100\nRacemark International LLC | Employment: 120"},
    {"id": "Q30", "category": "Supply Chain Risk & Resilience", "question": "Identify Georgia Tier 2/3 suppliers that are EV Relevant, classified as General Automotive, and provide materials to Battery Cell or Battery Pack companies, suggesting supply chain misalignment risk.", "golden": "There are 2 Georgia Tier 2/3 suppliers that are EV Relevant, classified as General Automotive, and provide materials to Battery Cell or Battery Pack companies, suggesting supply chain misalignment risk.\nDuckyang | OEMs: Hyundai Kia Rivian\nEnchem America Inc. | OEMs: Hyundai Kia Rivian"},
    {"id": "Q31", "category": "Supply Chain Risk & Resilience", "question": "Identify all Georgia-based Tier 1/2 automotive suppliers that maintain a diversified customer base (serving 'Multiple OEMs').", "golden": "There are 11 Georgia-based Tier 1/2 automotive suppliers that maintain a diversified customer base.  \nFouts Brothers Fire Equipment \nHitachi Astemo\nHwashin\nHyundai & LG Energy Solution (LGES)\nIMS Gear Georgia Inc.\nJAC Products Inc.\nHyundai Industrial Co.\nHyundai MOBIS (Georgia)\nHyundai Transys Georgia Seating Systems\nInalfa Roof Systems Inc.\nJefferson Southern Corp."},
    {"id": "Q32", "category": "Supply Chain Risk & Resilience", "question": "Identify Georgia companies in the Thermal Management or Power Electronics role with fewer than 200 employees - their small scale may limit surge production capacity.", "golden": "There are 3 Georgia companies in the Thermal Management or Power Electronics role with fewer than 200 employees.\n Freudenberg-NOK | Employment: 160 | Role: Thermal Management\n Hyundai Transys Georgia Powertrain | Employment: 130 | Role: Thermal Management\n Peerless-Winsmith Inc. | Employment: 160 | Role: Thermal Management"},
    {"id": "Q33", "category": "Supply Chain Risk & Resilience", "question": "Identify any EV-relevant Georgia companies classified as OEM Footprint or OEM Supply Chain?", "golden": "There are 24 Georgia\u2011based companies that are classified as OEM Footprint or OEM Supply Chain\n\nOEM Footprint\n\nTI Fluid Systems\nTie Down Engineering\nTN Americas Holding Inc.\nToyota Industries Group (TACG-TICA)\nTrenton Pressing\nYazaki North America\nYKK USA Inc.\nZF Gainesville LLC\n\nOEM Supply Chain:\n\nTrenton Pressing Inc.\nValeo\nVanguard National Trailer Corp.\nVernay\nVista Metals Corp.\nVoestalpine Automotive Body Parts Inc.\nVolvo Cars USA\nVolvo Group North America\nWabash National Corp.\nWheelabrator Group Inc.\nWIKA USA\nWoodbridge Foam Corp.\nWoory Industrial Co.\nYachiyo Manufacturing of America LLC\nYamaha Motor Manufacturing Corp.\nZF Gainesville LLC"},
    {"id": "Q34", "category": "Supply Chain Mapping & Visibility", "question": "Top 10 Georgia companies based on employment size that supply both General Automotive and EV-specific components, indicating transition readiness?", "golden": "Top 10 Georgia companies based on the employment size supply both General Automotive and have some EV relevance (Yes or Indirect): \n\nJTEKT North America Corp. | Employment: 860 \nKautex Inc. | Employment: 800 \nLark United Manufacturing Inc. | Employment: 700 \nFouts Brothers Fire Equipment | Employment: 630 \nArising Industries Inc. | Employment: 582 \nDinex Emissions Inc. | Employment: 500 \nLund International Inc. | Employment: 500 \nMack Trucks | Employment: 500 \nMando America Corp. | Employment: 460 \nACM Georgia LLC | Employment: 400"},
    {"id": "Q35", "category": "Product & Technology Trends", "question": "How many Georgia companies are now producing lithium-ion battery materials, cells, or electrolytes?", "golden": "There are 5 Georgia\u2011based companies currently producing lithium\u2011ion battery materials, battery cells, or battery electrolytes:\n  F&P Georgia Manufacturing [Tier 1/2] | Lithium-ion battery recycler & raw materials\n  Hitachi Astemo Americas Inc. [Tier 1/2] | Battery cells for electric mobility\n  Hollingsworth & Vose Co. [Tier 1/2] | Lithium-ion battery materials\n  Honda Development & Manufacturing [Tier 1/2] | Battery cells\n  IMMI [Tier 1/2] | Battery electrolyte"},
    {"id": "Q36", "category": "Product & Technology Trends", "question": "Which Georgia Tier 2/3 suppliers currently produce lightweight aluminum or composite materials and are growing their EV-specific customer base?", "golden": "There are 7 Tier 2/3 companies producing lightweight aluminum or composite materials and are growing their EV-specific customer base\nBridgestone Bandag | Product: Composite materials and specialty polymers for automotive components | EV Relevant: Indirect\nDinex Emissions Inc. | Product: Aluminum sheet for automotive bodies | EV Relevant: Indirect\nDown 2 Earth Trailers | Product: Standard and custom aluminum-lithium products for automotive customers | EV Relevant: Indirect\nEaton Corp. | Product: Aluminum products for various markets including automotive market | EV Relevant: Indirect\nEcoplastic America Corporation | Product: Aluminum sheet for automotive bodies | EV Relevant: Indirect\nErdrich USA Inc. | Product: Aluminum sheet for automotive bodies | EV Relevant: Indirect\nGlobal Powertrain Systems LLC | Product: Motor vehicle brake systems and parts Fabricates metal products, including coate | EV Relevant: Indirect"},
    {"id": "Q37", "category": "Product & Technology Trends", "question": "Identify Georgia companies whose product descriptions include 'high-voltage', 'DC-to-DC', 'inverter', or 'motor controller' \u2014 these signal EV powertrain electronics growth.", "golden": "There is only 1 Georgia company with 'high-voltage', 'DC-to-DC', 'inverter', or 'motor controller' product signals.\nGSC Steel Stamping LLC | Product: DC-to-DC converters"},
    {"id": "Q38", "category": "Product & Technology Trends", "question": "Which Georgia automotive companies employ over 1,000 workers but are currently categorized as only \"Indirectly Relevant\" to the EV sector, indicating a massive workforce pool ready for a strategic pivot?", "golden": "There are 3 Georgia automotive companies employ over 1,000 workers but are currently categorized as only \"Indirectly Relevant\" to the EV sector\nTCI Powder Coatings | Employment: 3000 | Role: Vehicle Assembly\nTeklas USA | Employment: 1800 | Role: Vehicle Assembly\nTextron Specialized Vehicles | Employment: 1100 | Role: Vehicle Assembly"},
    {"id": "Q39", "category": "Product & Technology Trends", "question": "Which three Georgia companies have the largest employment in EV thermal management, and how many employees does each have?", "golden": "The 4 companies with the largest thermal-management-related employment in Georgia are:\nZF Gainesville LLC: 17,500 employees\nNovelis Inc.: 200 employees\nFreudenberg-NOK: 160 employees\nPeerless-Winsmith Inc.: 160 employees"},
    {"id": "Q40", "category": "Product & Technology Trends", "question": "Which companies are involved in thermal-related products or services, and what roles and facility types are they associated with?", "golden": "There are 5 companies involved in thermal-related products or services, along with their associated EV supply chain roles and primary facility types are:\nFreudenberg-NOK | Role: Thermal Management | Facility Type: Manufacturing Plant\nHyundai Transys Georgia Powertrain |Role: Thermal Management | Facility Type: Manufacturing Plant\nNovelis Inc. | Role: Thermal Management | Facility Type: Manufacturing Plant\nPeerless-Winsmith Inc. | Role: Thermal Management | Facility Type: Manufacturing Plant\nZF Gainesville LLC | Role: EV thermal management and power electronics | Facility Type: Manufacturing"},
    {"id": "Q41", "category": "Product & Technology Trends", "question": "Which Tier 1/2 Georgia companies listed under General Automotive suggest a gradual evolution toward EV-related products or markets?", "golden": "There are 6 Tier 1/2 Georgia companies listed under General Automotive:\nFouts Brothers Fire Equipment [Tier 1/2] | Role: General Automotive\nHyundai Industrial Co. [Tier 1/2] | Role: General Automotive\nHyundai MOBIS (Georgia) [Tier 1/2] | Role: General Automotive\nHyundai Transys Georgia Seating Systems [Tier 1/2] | Role: General Automotive\nInalfa Roof Systems Inc. [Tier 1/2] | Role: General Automotive\nJefferson Southern Corp. [Tier 1/2] | Role: General Automotive"},
    {"id": "Q42", "category": "Product & Technology Trends", "question": "How is demand for thermal management solutions reflected in the number and employment size of Georgia Thermal Management suppliers?", "golden": "There are 4 companies for Thermal Management suppliers in Georgia\nEmployment Size: 650\nFreudenberg-NOK | Employment: 160\nHyundai Transys Georgia Powertrain | Employment: 130\nNovelis Inc. | Employment: 200\nPeerless-Winsmith Inc. | Employment: 160"},
    {"id": "Q43", "category": "Product & Technology Trends", "question": "Which Georgia companies are involved in battery recycling or second-life battery processing, reflecting the emerging circular economy trend?", "golden": "There are 3 Georgia companies involved in battery recycling or end-of-life processing:\nEnplas USA Inc. | Product: Recycler of lithium ion batteries\nEVCO Plastics | Product: Recycler of copper, precious metals, and non-ferrous materials\nF&P Georgia Manufacturing | Product: Lithium-ion battery recycler and raw materials provider"},
    {"id": "Q44", "category": "Product & Technology Trends", "question": "Which Georgia suppliers appear to play innovation-stage roles through research, development, or prototyping activity?", "golden": "There is only 1 Georgia company that clearly appears to be an innovation-stage roles through research, development, or prototyping activity.\nRacemark International LLC | Product: Manufacturing and R&D engine parts for EV"},
    {"id": "Q45", "category": "Product & Technology Trends", "question": "Which Georgia suppliers currently serving traditional OEMs are also linked to EV-native OEMs, showing dual-platform supply capability?", "golden": "There are 5 Georgia-based suppliers that exhibit dual-platform supply capability, as they are linked to both traditional OEMs such as Hyundai/Kia and the EV-native OEM Rivian.\nDuckyang\nGSC Steel Stamping LLC\nHyundai Transys Georgia Powertrain\nEnchem America Inc.\nRacemark International LLC"},
    {"id": "Q46", "category": "Site Selection & Expansion Planning", "question": "Identify Georgia areas that currently lack Battery Cell or Battery Pack suppliers but have existing Tier 1 general automotive infrastructure.", "golden": "There are 39 Georgia counties that have Tier 1 suppliers but no Battery Cell/Pack suppliers (using Updated Location).\nBaldwin County, Bibb County, Carroll County, Catoosa County, Chatham County, Chattahoochee County, Chattooga County, Clarke County, Clayton County, Cobb County, Dawson County, Dougherty County, Douglas County, Elbert County, Fayette County, Forsyth County, Franklin County, Fulton County, Grady County, Gwinnett County, Habersham County, Hall County, Haralson County, Henry County, Jackson County, Jones County, Lamar County, Lowndes County, Lumpkin County, Meriwether County, Morgan County, Newton County, Paulding County, Rabun County, Spalding County, Stephens County, Troup County, Walton County, Whitfield County"},
    {"id": "Q47", "category": "Site Selection & Expansion Planning", "question": "For a new Tier 1 battery thermal management company looking to locate in Georgia, which areas have the highest concentration of Materials-category suppliers that could support thermal management production?", "golden": "Covington, Henry County has 2 Materials\u2011category suppliers that could support battery thermal management production for a new Tier\u202f1 company in Georgia."},
    {"id": "Q48", "category": "Site Selection & Expansion Planning", "question": "How many Georgia areas have concentrated Manufacturing Plant facilities but no EV-specific production presence, indicating potential conversion-ready industrial sites?", "golden": "There are 100 Georgia areas that currently host Manufacturing Plant facilities but have no EV\u2011specific production presence, indicating a large pool of potentially conversion\u2011ready industrial sites.\nTop concentrations include:\nLaGrange, Troup County | 7 plants\nAtlanta, Fulton County | 6 plants\nWest Point, Troup County | 6 plants\nDublin, Laurens County | 4 plants\nSavannah, Chatham County | 4 plants\nGriffin, Spalding County | 3 plants\nMadison, Morgan County | 3 plants\nLavonia, Franklin County | 3 plants\nDuluth, Gwinnett County | 3 plants\nJefferson, Jackson County | 3 plants"},
    {"id": "Q49", "category": "Site Selection & Expansion Planning", "question": "For an international battery materials company seeking a Georgia location, which areas have existing chemical manufacturing infrastructure?", "golden": "There are 2 Georgia areas with existing chemical manufacturing infrastructure:\nCovington, Morgan County: Archer Aviation Inc.\nNorcross, Gwinnett County: Arising Industries Inc."},
    {"id": "Q50", "category": "Site Selection & Expansion Planning", "question": "Which Georgia areas have R&D facility types in the automotive sector, suggesting innovation infrastructure suitable for EV technology development centers?", "golden": "There is only 1 Georgia area with an R&D facility type in the automotive sector:\nGray, Jones County | Company: Racemark International LLC | Product: Manufacturing and R&D engine parts for EV"},
]

# ── Judge helpers ──────────────────────────────────────────────────────────────

def _clip(v: float) -> float:
    return max(0.0, min(1.0, float(v)))


def _trim(text: str, n: int) -> str:
    text = str(text or "")
    return text if len(text) <= n else text[: n // 2] + "\n...[trimmed]...\n" + text[-(n // 2):]


def _build_prompt(metric: str, question: str, golden: str, answer: str, context: str) -> str:
    ctx_block = (
        f"RETRIEVED CONTEXT:\n{_trim(context, 3000)}\n\n"
        if metric in {"faithfulness", "context_precision", "context_recall"} and context
        else ""
    )
    return (
        "You are an expert RAG evaluator.\n"
        f"Metric: {metric}\n"
        f"Definition: {DEFINITIONS[metric]}\n\n"
        f"QUESTION: {_trim(question, 500)}\n"
        f"GOLDEN ANSWER: {_trim(golden, 1500)}\n"
        f"GENERATED ANSWER: {_trim(answer, 1500)}\n"
        f"{ctx_block}"
        "TASK: Return ONLY valid JSON (no markdown):\n"
        '{"score": <float 0.0-1.0>, "reasoning": "<1-2 sentences>"}\n'
    )


def _parse(raw: str) -> dict:
    """Robust parser: handles markdown fences + partial JSON."""
    # Strip markdown fences
    raw = re.sub(r"```(?:json)?", "", raw).strip().strip("`").strip()
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        pass
    start, end = raw.find("{"), raw.rfind("}")
    if start != -1 and end > start:
        try:
            return json.loads(raw[start: end + 1])
        except json.JSONDecodeError:
            pass
    m = re.search(r'(?i)"?score"?\s*[:=]\s*([01](?:\.\d+)?)', raw)
    if m:
        return {"score": float(m.group(1)), "reasoning": raw[:200]}
    raise ValueError(f"Unparseable judge response: {raw[:150]}")


async def _call_judge(prompt: str, url: str, model: str) -> str:
    payload = {
        "model": model, "prompt": prompt, "stream": False,
        "format": "json",
        "options": {"temperature": 0, "num_predict": 200},
    }
    async with httpx.AsyncClient(timeout=60.0) as client:
        resp = await client.post(url, json=payload)
        resp.raise_for_status()
    body = resp.json()
    return str(body.get("response") or body.get("thinking") or "").strip()


async def score_metric(metric: str, q: str, golden: str, ans: str, ctx: str,
                       url: str, model: str) -> dict:
    prompt = _build_prompt(metric, q, golden, ans, ctx)
    for attempt in range(3):
        try:
            raw = await _call_judge(prompt, url, model)
            parsed = _parse(raw)
            s = _clip(float(parsed["score"]))
            r = str(parsed.get("reasoning", ""))[:300]
            return {"score": s, "reasoning": r}
        except Exception as exc:
            logger.warning("Judge attempt %d failed [%s]: %s", attempt + 1, metric, exc)
            await asyncio.sleep(1.0 * (attempt + 1))
    return {"score": 0.0, "reasoning": "evaluation_error_after_3_retries"}


async def evaluate_row(row: dict, url: str, model: str) -> dict:
    """Score all 5 metrics sequentially — critical for local Ollama (1 model at a time)."""
    scores: dict[str, dict] = {}
    for metric in DEFINITIONS:
        scores[metric] = await score_metric(
            metric, row["question"], row["golden"],
            row["answer"], row.get("context", ""), url, model,
        )

    final = _clip(sum(WEIGHTS[m] * scores[m]["score"] for m in DEFINITIONS))
    return {
        **row,
        "faithfulness":            scores["faithfulness"]["score"],
        "answer_relevancy":        scores["answer_relevancy"]["score"],
        "context_precision":       scores["context_precision"]["score"],
        "context_recall":          scores["context_recall"]["score"],
        "answer_correctness":      scores["answer_correctness"]["score"],
        "final_score":             final,
        "faithfulness_reason":     scores["faithfulness"]["reasoning"],
        "answer_relevancy_reason": scores["answer_relevancy"]["reasoning"],
        "context_precision_reason":scores["context_precision"]["reasoning"],
        "context_recall_reason":   scores["context_recall"]["reasoning"],
        "correctness_reason":      scores["answer_correctness"]["reasoning"],
    }


# ── Checkpoint helpers ─────────────────────────────────────────────────────────

def load_checkpoint() -> dict[str, dict]:
    """Load previously completed questions from progress file. Returns {id: row}."""
    done: dict[str, dict] = {}
    if PROGRESS_FILE.exists():
        for line in PROGRESS_FILE.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line:
                try:
                    row = json.loads(line)
                    done[row["id"]] = row
                except Exception:
                    pass
    return done


def save_checkpoint(row: dict) -> None:
    """Append one completed+scored row to the progress file."""
    PROGRESS_FILE.parent.mkdir(parents=True, exist_ok=True)
    with PROGRESS_FILE.open("a", encoding="utf-8") as f:
        f.write(json.dumps(row, ensure_ascii=False) + "\n")


def append_answers_md(row: dict) -> None:
    """Append human-readable entry to answers.md."""
    ANSWERS_MD.parent.mkdir(parents=True, exist_ok=True)
    with ANSWERS_MD.open("a", encoding="utf-8") as f:
        f.write(f"\n## {row['id']} [{row['category']}] — {row['elapsed_s']}s\n")
        f.write(f"**Q**: {row['question']}\n\n")
        f.write(f"**Golden**: {row['golden']}\n\n")
        f.write(f"**Generated**: {row['answer']}\n\n")
        f.write(
            f"**Scores**: faithfulness={row.get('faithfulness',0):.3f} | "
            f"relevancy={row.get('answer_relevancy',0):.3f} | "
            f"precision={row.get('context_precision',0):.3f} | "
            f"recall={row.get('context_recall',0):.3f} | "
            f"correctness={row.get('answer_correctness',0):.3f} | "
            f"**final={row.get('final_score',0):.3f}**\n\n"
            "---\n"
        )


# ── Pipeline runner ────────────────────────────────────────────────────────────

def run_one(agent: EVAgent, q: dict) -> dict:
    t0 = time.monotonic()
    try:
        result = agent.ask(q["question"])
        return {
            "id":       q["id"],
            "category": q["category"],
            "question": q["question"],
            "golden":   q["golden"],
            "answer":   result["answer"],
            "context":  result.get("retrieved_context", ""),
            "elapsed_s": round(time.monotonic() - t0, 1),
            "path":     result["entities"].get(
                "retrieval_source",
                "cypher" if result["entities"].get("cypher_used") else "sql",
            ),
        }
    except Exception as exc:
        logger.error("Pipeline error on %s: %s", q["id"], exc)
        return {
            "id":       q["id"], "category": q["category"],
            "question": q["question"], "golden": q["golden"],
            "answer":   f"[PIPELINE ERROR: {exc}]", "context": "",
            "elapsed_s": round(time.monotonic() - t0, 1), "path": "error",
        }


# ── Excel report ───────────────────────────────────────────────────────────────

HEADERS = [
    "Q_ID","Category","Question","Golden_Answer","Generated_Answer","Path","Elapsed_s",
    "Faithfulness","Answer_Relevancy","Context_Precision","Context_Recall",
    "Answer_Correctness","Final_Score",
    "Faithfulness_Reason","Relevancy_Reason","Precision_Reason","Recall_Reason","Correctness_Reason",
]


def _fill(v: float) -> PatternFill:
    if v >= 0.7: return PatternFill(fill_type="solid", fgColor="C6EFCE")
    if v >= 0.5: return PatternFill(fill_type="solid", fgColor="FFF2CC")
    return PatternFill(fill_type="solid", fgColor="F4CCCC")


def build_report(results: list[dict], out: Path) -> None:
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Results"
    hfill = PatternFill(fill_type="solid", fgColor="1F4E78")
    wfont = Font(color="FFFFFF", bold=True)
    ws.append(HEADERS)
    for c in ws[1]:
        c.fill = hfill; c.font = wfont
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    lb = PatternFill(fill_type="solid", fgColor="D9EAF7")
    wh = PatternFill(fill_type="solid", fgColor="FFFFFF")
    for i, r in enumerate(results, 2):
        ws.append([
            r.get("id",""), r.get("category",""), r.get("question",""),
            r.get("golden",""), r.get("answer",""), r.get("path",""), r.get("elapsed_s",0),
            r.get("faithfulness",0), r.get("answer_relevancy",0),
            r.get("context_precision",0), r.get("context_recall",0),
            r.get("answer_correctness",0), r.get("final_score",0),
            r.get("faithfulness_reason",""), r.get("answer_relevancy_reason",""),
            r.get("context_precision_reason",""), r.get("context_recall_reason",""),
            r.get("correctness_reason",""),
        ])
        ws.row_dimensions[i].height = 70
        rf = lb if i % 2 == 0 else wh
        for col in range(1, len(HEADERS)+1):
            ws.cell(i, col).fill = rf
            ws.cell(i, col).alignment = Alignment(vertical="top", wrap_text=True)
        for col in range(8, 14):
            cell = ws.cell(i, col)
            try:
                cell.fill = _fill(float(cell.value or 0))
                cell.number_format = "0.0000"
            except (TypeError, ValueError):
                pass

    ws.freeze_panes = "A2"
    for col, w in {"A":8,"B":16,"C":52,"D":55,"E":75,"F":8,"G":8}.items():
        ws.column_dimensions[col].width = w

    # Summary sheet
    ss = wb.create_sheet("Summary")
    ss.append(["Metric","Mean","Std Dev","Weight"])
    for c in ss[1]:
        c.fill = hfill; c.font = wfont
    for mk in list(DEFINITIONS) + ["final_score"]:
        vals = [float(r.get(mk, 0) or 0) for r in results]
        mu = mean(vals) if vals else 0.0
        sd = pstdev(vals) if len(vals) > 1 else 0.0
        ss.append([mk, round(mu, 4), round(sd, 4), WEIGHTS.get(mk, "-")])
        c = ss.cell(ss.max_row, 2)
        c.fill = _fill(mu); c.number_format = "0.0000"

    chart = BarChart()
    chart.title = "Phase 4 RAGAS Scores"
    chart.y_axis.title = "Score"
    data = Reference(ss, min_col=2, min_row=1, max_row=ss.max_row)
    cats = Reference(ss, min_col=1, min_row=2, max_row=ss.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 9; chart.width = 16
    ss.add_chart(chart, "F2")

    # Timing sheet
    ts = wb.create_sheet("Timing")
    ts.append(["Q_ID","Category","Elapsed_s","Path","Final_Score"])
    for c in ts[1]:
        c.fill = hfill; c.font = wfont
    for r in results:
        ts.append([r.get("id",""), r.get("category",""), r.get("elapsed_s",0),
                   r.get("path",""), r.get("final_score",0)])

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    logger.info("Report saved: %s", out)


# ── Console summary ────────────────────────────────────────────────────────────

def print_summary(results: list[dict]) -> None:
    print(f"\n{'='*68}")
    print("  RAGAS EVALUATION SUMMARY — Phase 4 Georgia EV Intelligence")
    print(f"{'='*68}")
    print(f"  Questions : {len(results)}  |  Judge: qwen2.5:7b (local Ollama)")
    print(f"{'-'*68}")
    for mk in list(DEFINITIONS) + ["final_score"]:
        vals = [float(r.get(mk,0) or 0) for r in results]
        mu = mean(vals) if vals else 0.0
        flag = "✅" if mu >= 0.7 else ("⚠️ " if mu >= 0.5 else "❌")
        w = f"w={WEIGHTS.get(mk,'—')}" if mk != "final_score" else "WEIGHTED"
        print(f"  {flag} {mk:<24} {mu:.4f}  {w}")
    print(f"{'='*68}")
    print(f"\n  Per-question:")
    print(f"  {'ID':<5} {'Category':<14} {'Time':>7} {'Score':>7}  {'Path':<6}  Preview")
    print("  " + "-"*72)
    for r in results:
        fs = float(r.get("final_score",0) or 0)
        flag = "✅" if fs >= 0.7 else ("⚠️" if fs >= 0.5 else "❌")
        prev = (r.get("answer","") or "")[:50].replace("\n"," ")
        print(f"  {r['id']:<5} {r['category']:<14} {r.get('elapsed_s',0):>6.1f}s"
              f" {fs:>6.3f} {flag} {r.get('path',''):6}  {prev}")
    ts = [r.get("elapsed_s",0) for r in results if r.get("path") != "error"]
    if ts:
        print(f"\n  ⏱  Avg {mean(ts):.1f}s | Max {max(ts):.1f}s | Min {min(ts):.1f}s\n")


# ── Main ──────────────────────────────────────────────────────────────────────

async def main(questions: list[dict], out_path: Path, resume: bool) -> None:
    cfg       = Config.get()
    judge_url = f"{cfg.ollama_base_url}/api/generate"
    model     = cfg.ollama_llm_model   # qwen2.5:7b — no extra RAM

    # Load checkpoint
    done = load_checkpoint() if resume else {}
    if done:
        print(f"  Resuming — {len(done)} questions already done, skipping them.")
        if not resume and PROGRESS_FILE.exists():
            PROGRESS_FILE.unlink()  # fresh run: clear checkpoint

    agent  = EVAgent()
    total  = len(questions)
    scored: list[dict] = list(done.values())

    for i, q in enumerate(questions, 1):
        if q["id"] in done:
            print(f"  [{i}/{total}] {q['id']} — skipped (checkpoint)")
            continue

        print(f"\n  [{i}/{total}] {q['id']} [{q['category']}]")
        print(f"  Q: {q['question'][:80]}")

        # Step 1: Pipeline
        t0  = time.monotonic()
        row = run_one(agent, q)
        print(f"  → Answer ({row['elapsed_s']}s, {row['path']}): {row['answer'][:80]}")

        # Step 2: RAGAS scoring
        print(f"  → Scoring 5 metrics...")
        scored_row = await evaluate_row(row, judge_url, model)
        answer_correctness = float(scored_row.get("answer_correctness", 0) or 0)
        flag = (
            "✅" if answer_correctness >= 0.7
            else ("⚠️" if answer_correctness >= 0.5 else "❌")
        )
        print(f"  → Answer correctness: {answer_correctness:.3f} {flag}")

        # Checkpoint immediately
        save_checkpoint(scored_row)
        append_answers_md(scored_row)
        scored.append(scored_row)

    # Build report
    print_summary(scored)
    build_report(scored, out_path)
    print(f"  📄 Excel : {out_path}")
    print(f"  📝 MD    : {ANSWERS_MD}")
    print(f"  💾 JSONL : {PROGRESS_FILE}\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--questions", type=int, default=7)
    parser.add_argument("--resume",    action="store_true",
                        help="Resume from checkpoint (skip already-scored questions)")
    parser.add_argument("--out", type=str,
                        default=str(_default_report_path()))
    args = parser.parse_args()

    qs = FIFTY_QUESTIONS[:args.questions]
    asyncio.run(main(qs, Path(args.out), args.resume))
