"""
scripts/export_questions.py
Reads the human-validated 50 questions Excel and prints them as Python list
for copy-pasting into run_ragas_eval.py
Run: venv\Scripts\python scripts\export_questions.py
"""
import json, sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

CANDIDATES = [
    Path(r"C:\Users\sande\OneDrive\Desktop\EV_Georgia_supplyChain\ev_data_LLM_comparsions\data\Human validated 50 questions.xlsx"),
    Path(r"C:\Users\sande\OneDrive\Desktop\EV_Georgia_supplyChain\ev_data_LLM_comparsions\Human validated 50 questions.xlsx"),
]

for path in CANDIDATES:
    if path.exists():
        break
else:
    print("ERROR: Could not find the Excel file. Tried:")
    for p in CANDIDATES:
        print(f"  {p}")
    sys.exit(1)

wb = openpyxl.load_workbook(path)
ws = wb.active
headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
print(f"Columns found: {headers}\n")

rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    d = {headers[i]: (str(v).strip() if v else "") for i, v in enumerate(row) if i < len(headers)}
    if any(d.values()):
        rows.append(d)

print(f"Total questions: {len(rows)}\n")
print("─" * 60)
print("# Paste this into SMOKE_QUESTIONS in run_ragas_eval.py")
print("─" * 60)
print("FIFTY_QUESTIONS = [")
for i, r in enumerate(rows, 1):
    # Try common column name variants
    q_col      = next((r[k] for k in ["Question","question","QUESTION"] if k in r and r[k]), "")
    golden_col = next((r[k] for k in ["Golden Answer","golden_answer","Golden","Answer","ANSWER","Expected Answer","Expected_Answer","golden"] if k in r and r[k]), "")
    cat_col    = next((r[k] for k in ["Category","category","Use Case","Use_Case","Type","type"] if k in r and r[k]), "GENERAL")
    id_col     = next((r[k] for k in ["ID","id","Q_ID","Question_ID","No","No.","#"] if k in r and r[k]), f"Q{i}")
    print(f'    {{')
    print(f'        "id": "{id_col}",')
    print(f'        "category": "{cat_col}",')
    print(f'        "question": {json.dumps(q_col)},')
    print(f'        "golden":   {json.dumps(golden_col)},')
    print(f'    }},')
print("]")
print("\n# Also saving to: outputs/questions_export.json")

out = Path("outputs/questions_export.json")
out.parent.mkdir(parents=True, exist_ok=True)
out.write_text(json.dumps(rows, indent=2, ensure_ascii=False), encoding="utf-8")
print(f"Saved {len(rows)} rows to {out}")
